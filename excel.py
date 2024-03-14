import openpyxl



def write_dict_array_to_excel(dict_array, excel_file_path):
	# 创建一个新的 Excel 工作簿
	wb = openOrCreateExcel(excel_file_path)
	# 获取默认的工作表
	ws = wb.active
	if dict_array:
		# 写入列名
		headers = list(dict_array[0].keys())
		for col_idx, header in enumerate(headers, start=1):
			ws.cell(row=1, column=col_idx, value=header)
	    # 写入数据
		for row_idx, data in enumerate(dict_array, start=2):
			for col_idx, header in enumerate(headers, start=1):
				ws.cell(row=row_idx, column=col_idx, value=data.get(header))
		# 保存 Excel 文件
		wb.save(excel_file_path)
		print(f"Data has been written to '{excel_file_path}'")

def clear_excel(excel_file_path):
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook(excel_file_path)
        print(f"Excel file '{excel_file_path}' has been opened successfully.")
        
        # 遍历每个工作表并清空单元格值
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        
        # 保存 Excel 文件
        wb.save(excel_file_path)
        print(f"Excel file '{excel_file_path}' has been cleared successfully.")
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file_path}' not found.")

def openOrCreateExcel(excel_file_path):
	try:
		# 打开 Excel 文件
		openpyxl.load_workbook(excel_file_path)
		clear_excel(excel_file_path)
		return openpyxl.load_workbook(excel_file_path)
	except FileNotFoundError:
		# 创建一个新的 Excel 工作簿
		return openpyxl.Workbook()


